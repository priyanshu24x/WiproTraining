from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook


def write_excel(filename)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["name":"age"])
    sheet.append((["john dow", "30"]))
    sheet.append((["jane sone", "35"]))
    workbook.save(filename)

def read_excel(filename):
    workbook = load_workbook(filename)
    sheet = workbook.active

def delete_excel(filename)
    pass

filename = "data.xlsx"
write_excel(filename)
print("data read from excel file: ")
read_excel(filename)
delete_excel(filename)

